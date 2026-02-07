import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill

def generar_excel_gastos(payload):
    data = payload.get('movimientos', [])
    titulo_ciclo = payload.get('titulo', 'REPORTE DE MOVIMIENTOS')
    
    df = pd.DataFrame(data)

    if df.empty:
        # Manejo simple si no hay datos
        output = io.BytesIO()
        pd.DataFrame([["Sin datos"]]).to_excel(output, index=False, header=False)
        output.seek(0)
        return output

    # 1. Lógica de Negativos (Gastos y Ahorros restan del saldo)
    if 'type' in df.columns and 'amount' in df.columns:
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
        # Guardamos los montos originales para sumatorias de totales antes de aplicar el signo
        df['monto_puro'] = df['amount']
        # Aplicamos negativo para el listado detallado
        df.loc[df['type'].isin(['gasto', 'ahorro']), 'amount'] = df['amount'] * -1

    # 2. Cálculos de Totales (Usando montos puros)
    totales = df.groupby('type')['monto_puro'].sum()
    total_ingresos = totales.get('ingreso', 0)
    total_gastos = totales.get('gasto', 0)
    total_ahorro = totales.get('ahorro', 0)
    saldo_final = total_ingresos - total_gastos - total_ahorro

    # 3. Mapeo y Limpieza para la tabla
    columnas_map = {'created_at': 'Fecha', 'description_user': 'Descripción', 
                    'category': 'Categoría', 'amount': 'Monto', 'type': 'Tipo'}
    
    tipos_originales = df['type'].tolist()
    columnas_existentes = [col for col in columnas_map.keys() if col in df.columns]
    df_export = df[columnas_existentes].rename(columns=columnas_map)

    if 'Fecha' in df_export.columns:
        df_export['Fecha'] = pd.to_datetime(df_export['Fecha']).dt.strftime('%d/%m/%Y %H:%M')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Reporte', startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets['Reporte']

        # --- TITULO ---
        # --- INSERTAR TÍTULO PREMIUM ---
        worksheet.merge_cells('A1:E1')
        titulo_celda = worksheet['A1']
        titulo_celda.value = titulo_ciclo

        # Usamos el color del texto de tu logo (un cian/celeste vibrante)
        # Hex: 06B6D4 (Cian) o 3B82F6 (Azul brillante)
        color_primario = "FFFFFF" 
        fondo_oscuro = "374151"

        titulo_celda.font = Font(bold=True, size=16, color=color_primario) # Letras Cian
        titulo_celda.alignment = Alignment(horizontal="center", vertical="center")
        # Fondo muy oscuro para que resalte el Cian
        titulo_celda.fill = PatternFill(start_color=fondo_oscuro, end_color=fondo_oscuro, fill_type="solid")

        # Aumentamos un poco el alto de la fila del título para que respire
        worksheet.row_dimensions[1].height = 30

        # --- ESTILO PARA EL ENCABEZADO DE LA TABLA (Fila 3) ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Gris azulado

        for col_num in range(1, 6):
            cell = worksheet.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # --- ANCHOS DE COLUMNA ---
        for idx, col in enumerate(df_export.columns):
            max_len = max(df_export[col].astype(str).map(len).max(), len(col)) + 4
            worksheet.column_dimensions[chr(65 + idx)].width = max_len

        # --- COLORES EN LISTADO ---
        monto_col_idx = df_export.columns.get_loc("Monto") + 1
        monto_col_letter = chr(64 + monto_col_idx)

        for row_idx, tipo in enumerate(tipos_originales, start=4): 
            cell = worksheet[f"{monto_col_letter}{row_idx}"]
            if tipo == 'gasto': cell.font = Font(color="E11D48", bold=True)
            elif tipo == 'ingreso': cell.font = Font(color="15803D", bold=True)
            elif tipo == 'ahorro': cell.font = Font(color="1D4ED8", bold=True)
            cell.number_format = '#,##0'

        # --- SECCIÓN DE TOTALES AL FINAL ---
        start_totals = len(df_export) + 5  # 2 filas después de la tabla
        
        # 1. Etiquetas de desglose (Sin combinar aún)
        etiquetas = [
            ("TOTAL INGRESOS (+):", total_ingresos, "15803D"),
            ("TOTAL GASTOS (-):", total_gastos * -1, "E11D48"),
            ("TOTAL AHORRO (-):", total_ahorro * -1, "1D4ED8"),
        ]

        for i, (label, value, color) in enumerate(etiquetas):
            row = start_totals + i
            # Etiqueta en columna C
            cell_label = worksheet[f"C{row}"]
            cell_label.value = label
            cell_label.alignment = Alignment(horizontal="right")
            
            # Valor en columna D
            cell_val = worksheet[f"D{row}"]
            cell_val.value = value
            cell_val.font = Font(bold=True, color=color)
            cell_val.number_format = '#,##0'

        # 2. EL SALDO NETO (Destacado al final de los totales)
        # Lo ponemos una fila más abajo del último desglose
        saldo_row = start_totals + len(etiquetas) + 1
        
        # Combinamos C y D para que el "Botón" de saldo sea grande
        worksheet.merge_cells(f'C{saldo_row}:D{saldo_row}')
        
        saldo_cell = worksheet[f"C{saldo_row}"]
        saldo_cell.value = f"SALDO NETO: {saldo_final:,.0f}"
        saldo_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo Premium para el Saldo
        fondo_saldo = "374151" if saldo_final >= 0 else "E11D48"
        saldo_cell.font = Font(bold=True, size=14, color="FFFFFF")
        saldo_cell.fill = PatternFill(start_color=fondo_saldo, end_color=fondo_saldo, fill_type="solid")
        
        # Le damos un poco más de altura a esa fila para que luzca
        worksheet.row_dimensions[saldo_row].height = 25

    output.seek(0)
    return output